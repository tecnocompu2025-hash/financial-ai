from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

TRANSLATIONS = {
    "income_total": "Ingresos Totales",
    "expense_total": "Gastos Totales",
    "passive_income_total": "Ingresos Pasivos",
    "cash_flow": "Flujo de Caja",
    "asset_total": "Total Activos",
    "liability_total": "Total Pasivos",
    "net_worth": "Patrimonio Neto",
    "debt_ratio": "Ratio de Deuda (%)",
    "emergency_fund_available": "Fondo de Emergencia",
    "essential_monthly_expenses": "Gastos Esenciales (Mes)",
    "emergency_months": "Meses de Supervivencia",
    "financial_freedom_percentage": "Libertad Financiera (%)",
    "productive_assets": "Activos Productivos"
}

def translate(key: str) -> str:
    return TRANSLATIONS.get(key, key.replace("_", " ").title())

class ReportExportService:
    @staticmethod
    def _apply_excel_header(ws, headers):
        ws.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="0F172A") # slate-950
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_num)].width = 20

    @staticmethod
    def excel(report, symbol="$ "):
        workbook = Workbook()
        
        # 1. Resumen
        summary = workbook.active
        summary.title = "Resumen Financiero"
        ReportExportService._apply_excel_header(summary, ["Indicador", "Monto"])
        for key, value in report.summary.model_dump().items():
            row = summary.append([translate(key), value])
            summary.cell(row=summary.max_row, column=2).number_format = f'"{symbol}"#,##0.00'

        # 2. Movimientos
        transactions = workbook.create_sheet("Movimientos")
        ReportExportService._apply_excel_header(transactions, ["Fecha", "Tipo", "Categoría", "Detalle", "Monto"])
        for item in report.transactions:
            transactions.append([item.date.isoformat(), "Ingreso" if item.record_type == "income" else "Gasto", item.category, item.name, item.amount])
            transactions.cell(row=transactions.max_row, column=5).number_format = f'"{symbol}"#,##0.00'
        transactions.column_dimensions['D'].width = 30

        # 3. Evolución
        evolution = workbook.create_sheet("Evolución Mensual")
        ReportExportService._apply_excel_header(evolution, ["Mes", "Ingresos", "Gastos", "Balance"])
        for item in report.monthly_evolution:
            evolution.append([item.month, item.income, item.expense, item.balance])
            evolution.cell(row=evolution.max_row, column=2).number_format = f'"{symbol}"#,##0.00'
            evolution.cell(row=evolution.max_row, column=3).number_format = f'"{symbol}"#,##0.00'
            evolution.cell(row=evolution.max_row, column=4).number_format = f'"{symbol}"#,##0.00'

        # 4. Salud
        health = workbook.create_sheet("Salud Financiera")
        ReportExportService._apply_excel_header(health, ["Indicador", "Valor"])
        for key, value in report.financial_health.items():
            health.append([translate(key), value])
            if "percentage" not in key and "months" not in key:
                health.cell(row=health.max_row, column=2).number_format = f'"{symbol}"#,##0.00'

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    @staticmethod
    def pdf(report, symbol="$ "):
        stream = BytesIO()
        doc = SimpleDocTemplate(stream, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(name='TitleStyle', parent=styles['Heading1'], fontSize=20, spaceAfter=20, textColor=colors.HexColor("#0F172A"))
        heading_style = ParagraphStyle(name='HeadingStyle', parent=styles['Heading2'], fontSize=14, spaceAfter=10, spaceBefore=20, textColor=colors.HexColor("#0891B2"))
        normal_style = styles['Normal']
        
        elements = []
        elements.append(Paragraph("Reporte Financiero Consolidado", title_style))
        elements.append(Paragraph("Resumen Financiero", heading_style))

        # Tabla de resumen
        summary_data = [["Indicador", "Monto"]]
        for key, value in report.summary.model_dump().items():
            summary_data.append([translate(key), f"{symbol}{value:,.2f}"])
            
        t = Table(summary_data, colWidths=[250, 150])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#E2E8F0")),
        ]))
        elements.append(t)
        
        # Tabla de Salud Financiera
        elements.append(Paragraph("Salud Financiera", heading_style))
        health_data = [["Indicador", "Valor"]]
        for key, value in report.financial_health.items():
            fmt_val = f"{value:,.2f}" if "percentage" in key or "months" in key else f"{symbol}{value:,.2f}"
            health_data.append([translate(key), fmt_val])
            
        t2 = Table(health_data, colWidths=[250, 150])
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#E2E8F0")),
        ]))
        elements.append(t2)

        # Recomendaciones
        elements.append(Paragraph("Recomendaciones de AI", heading_style))
        for rec in report.recommendations:
            elements.append(Paragraph(f"• {rec}", normal_style))
            elements.append(Spacer(1, 6))

        doc.build(elements)
        stream.seek(0)
        return stream
